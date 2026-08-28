"""FastAPI app: REST API + dashboard for the NSE EOD Momentum strategy."""
from __future__ import annotations

from dataclasses import asdict
from pathlib import Path

from fastapi import Body, FastAPI, HTTPException
from fastapi.responses import FileResponse, RedirectResponse, Response

from .config import (DATA_DIR, EODConfig, EnvSettings, load_config, save_config)
from .engine import EODEngine, TRADES_FILE
from .upstox_client import UpstoxClient, UpstoxError

FRONTEND = Path(__file__).resolve().parent.parent / "frontend" / "index.html"

env = EnvSettings()
client = UpstoxClient(env.api_key, env.api_secret, env.redirect_uri)
engine = EODEngine(client, load_config())

app = FastAPI(title="NSE EOD Momentum Options Strategy (Upstox)")

_INT = {"top_n", "lots", "otm_strikes", "shortlist_size", "universe_limit", "hedge_itm_strikes"}
_NUM = {"tp_pct", "sl_pct", "tick_size", "min_oi", "max_spread_pct", "capital", "charges_per_trade"}
_STR = {"mode", "scan_time", "refresh_time", "dispatch_time"}
_BOOL = {"use_gtt", "hedge_itm", "require_hedge"}


@app.on_event("startup")
async def _startup():
    engine.ensure_loop()


@app.get("/")
async def index():
    return FileResponse(FRONTEND)


# ---------------- auth ----------------
@app.get("/api/auth/login")
async def auth_login():
    if not env.api_key or not env.api_secret:
        raise HTTPException(400, "Set UPSTOX_API_KEY and UPSTOX_API_SECRET in .env first")
    return RedirectResponse(client.login_url())


@app.get("/api/auth/callback")
async def auth_callback(code: str = ""):
    if not code:
        raise HTTPException(400, "Missing authorization code")
    try:
        await client.exchange_code(code)
    except UpstoxError as e:
        raise HTTPException(400, str(e))
    engine.log("auth", "Connected to Upstox")
    return RedirectResponse("/")


@app.post("/api/auth/token")
async def auth_token(body: dict = Body(...)):
    token = (body.get("access_token") or "").strip()
    if not token:
        raise HTTPException(400, "access_token required")
    client.set_token(token)
    engine.log("auth", "Access token set manually")
    return await auth_status()


@app.post("/api/auth/logout")
async def auth_logout():
    client.clear_token()
    engine.running = False
    engine.log("auth", "Disconnected from Upstox")
    return {"ok": True}


@app.get("/api/auth/status")
async def auth_status():
    info = {"configured": bool(env.api_key and env.api_secret),
            "token_present": bool(client.access_token), "profile": None}
    if client.access_token:
        try:
            p = await client.profile()
            info["profile"] = {"name": p.get("user_name"), "user_id": p.get("user_id"),
                               "broker": p.get("broker")}
        except UpstoxError as e:
            info["error"] = str(e)
    return info


# ---------------- status / control ----------------
@app.get("/api/status")
async def status():
    return engine.status()


@app.post("/api/start")
async def start():
    try:
        engine.start()
    except UpstoxError as e:
        raise HTTPException(400, str(e))
    return {"running": True}


@app.post("/api/stop")
async def stop():
    engine.stop()
    return {"running": False}


@app.post("/api/scan")
async def scan():
    """Run the scan + refresh now (does not place orders)."""
    if not client.access_token:
        raise HTTPException(400, "Connect to Upstox first")
    try:
        cands = await engine.scan_now()
    except UpstoxError as e:
        raise HTTPException(400, str(e))
    return {"candidates": cands, "shortlist": engine.scan_result}


@app.post("/api/dispatch")
async def dispatch():
    """Force the sell dispatch now (manual trigger; normally fires at 3:25 PM)."""
    if not client.access_token:
        raise HTTPException(400, "Connect to Upstox first")
    if not engine.candidates:
        raise HTTPException(400, "No candidates — run a scan first")
    from datetime import datetime
    from .engine import TZ
    await engine._dispatch(datetime.now(TZ))
    return {"positions": [p for p in engine.positions if p["status"] == "open"]}


@app.post("/api/exit")
async def manual_exit(body: dict = Body(...)):
    pid = body.get("id")
    ok = await engine.manual_exit(pid)
    if not ok:
        raise HTTPException(400, "Position not found or already closed")
    return {"ok": True}


@app.post("/api/clear")
async def clear_position(body: dict = Body(...)):
    """Remove a position from tracking without placing any market order."""
    ok = await engine.clear_position(body.get("id"))
    if not ok:
        raise HTTPException(400, "Position not found or already closed")
    return {"ok": True}


@app.get("/api/config")
async def get_config():
    return asdict(engine.cfg)


@app.post("/api/config")
async def set_config(body: dict = Body(...)):
    cur = asdict(engine.cfg)
    for k, v in body.items():
        if k in _INT:
            cur[k] = int(v)
        elif k in _NUM:
            cur[k] = float(v)
        elif k in _BOOL:
            cur[k] = bool(v)
        elif k in _STR:
            cur[k] = v
    try:
        cfg = EODConfig(**cur)
        cfg.validate()
    except (ValueError, TypeError) as e:
        raise HTTPException(400, str(e))
    engine.cfg = cfg
    save_config(cfg)
    engine.log("config", "Settings updated")
    return asdict(cfg)


@app.get("/api/trades")
async def trades():
    return engine.trades


EXPORT_COLUMNS = [
    ("entry_day", "Entry Day"), ("exit_day", "Exit Day"), ("mode", "Mode"),
    ("symbol", "Stock"), ("bias", "Bias"), ("side", "Option"),
    ("option_symbol", "Contract"), ("strike", "Strike"), ("expiry", "Expiry"),
    ("qty", "Qty"), ("spot_at_entry", "Spot @ Entry"),
    ("sell_price", "Sell ₹"), ("exit_price", "Exit ₹"), ("short_net_rs", "Short Net ₹"),
    ("hedge_symbol", "Hedge (ITM)"), ("hedge_side", "Hedge Side"),
    ("hedge_strike", "Hedge Strike"), ("buy_price", "Hedge Buy ₹"),
    ("hedge_exit", "Hedge Exit ₹"), ("hedge_net_rs", "Hedge Net ₹"),
    ("gross_rs", "Gross ₹"), ("charges_rs", "Charges ₹"), ("net_rs", "Net ₹"),
    ("entry_time", "Entry Time"), ("exit_time", "Exit Time"), ("reason", "Exit Reason"),
]


def _all_trades() -> list[dict]:
    import json
    out = []
    if TRADES_FILE.exists():
        for line in TRADES_FILE.read_text().splitlines():
            try:
                out.append(json.loads(line))
            except Exception:
                continue
    return out


@app.get("/api/trades.xlsx")
async def trades_xlsx(scope: str = "all"):
    rows = engine.trades if scope == "today" else _all_trades()
    keys = [k for k, _ in EXPORT_COLUMNS]
    headers = [h for _, h in EXPORT_COLUMNS]
    fname = f"nse-eod-trades-{'today' if scope == 'today' else 'all'}.xlsx"
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook()
        ws = wb.active
        ws.title = "EOD Trades"
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for t in rows:
            ws.append([t.get(k) for k in keys])
        ws.freeze_panes = "A2"
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(11, len(h) + 2)
        buf = BytesIO()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'})
    except ImportError:
        import csv
        import io
        sio = io.StringIO()
        w = csv.writer(sio)
        w.writerow(headers)
        for t in rows:
            w.writerow([t.get(k) for k in keys])
        return Response(sio.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="{fname[:-5]}.csv"'})


def main() -> None:
    import uvicorn
    uvicorn.run("app.main:app", host=env.host, port=env.port, reload=False)


if __name__ == "__main__":
    main()
